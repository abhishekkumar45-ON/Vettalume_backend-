"""Admin content portal API. Every route in this router sits behind require_admin (router-level
dependency), so an unauthenticated caller gets 401 and an authenticated non-admin gets 403. This is
the "no outsiders" perimeter for authoring the syllabus graph and the item bank.

What it exposes:
  * syllabus read/write  — exams, sections, topics, concepts, prerequisite edges
  * item lifecycle       — list (with answers), create, edit, approve, retire, delete, bulk xlsx
  * admin management     — list / grant / revoke admins

Bulk question volume still goes through the same importer the /ingest path uses; the portal is for
the things a spreadsheet is bad at (defining the graph, approving/QC, spot-editing single items).
"""
from __future__ import annotations

import io
from typing import Any, Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from pydantic import BaseModel, ConfigDict, Field
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from .. import models
from ..deps import get_db
from ..schemas import IngestReport, ItemIn
from ..services import ingestion, question_bank
from ..services.admin_auth import grant_admin, require_admin, revoke_admin

router = APIRouter(prefix="/admin", tags=["admin"], dependencies=[Depends(require_admin)])


# ───────────────────────── whoami ─────────────────────────
@router.get("/me")
def admin_me(admin: models.Account = Depends(require_admin)) -> dict:
    return {"account_id": str(admin.id), "email": admin.email,
            "display_name": admin.display_name, "is_admin": True}


# ───────────────────────── syllabus: read ─────────────────────────
@router.get("/exams")
def list_exams(db: Session = Depends(get_db)) -> list[dict]:
    return [{"code": e.code, "name": e.name} for e in db.scalars(select(models.Exam)).all()]


@router.get("/syllabus")
def syllabus(exam: str, db: Session = Depends(get_db)) -> dict:
    ex = db.get(models.Exam, exam)
    if ex is None:
        raise HTTPException(404, f"no exam {exam!r}")
    sections = db.scalars(select(models.Section).where(models.Section.exam_code == exam)).all()
    nodes = db.scalars(select(models.KnowledgeNode).where(models.KnowledgeNode.exam_code == exam)).all()
    prereqs = db.scalars(select(models.PrereqEdge)).all()
    items = db.scalars(select(models.Item).where(models.Item.exam_code == exam)).all()
    sec_by_id = {s.id: s for s in sections}
    item_count: dict[str, int] = {}
    for it in items:
        item_count[it.concept_node_id] = item_count.get(it.concept_node_id, 0) + 1
    pre_by_node: dict[str, list[str]] = {}
    for e in prereqs:
        pre_by_node.setdefault(e.node_id, []).append(e.prereq_node_id)
    return {
        "exam": {"code": ex.code, "name": ex.name},
        "sections": [{"id": str(s.id), "key": s.key, "name": s.name} for s in sections],
        "nodes": [{
            "id": n.id, "name": n.name, "kind": n.kind, "parent_id": n.parent_id,
            "section": sec_by_id[n.section_id].key if n.section_id in sec_by_id else None,
            "item_count": item_count.get(n.id, 0),
            "prereqs": pre_by_node.get(n.id, []),
        } for n in nodes],
    }


# ───────────────────────── syllabus: write ─────────────────────────
class ExamIn(BaseModel):
    code: str
    name: str


@router.post("/exams")
def create_exam(body: ExamIn, db: Session = Depends(get_db)) -> dict:
    if db.get(models.Exam, body.code):
        raise HTTPException(409, f"exam {body.code!r} already exists")
    db.add(models.Exam(code=body.code, name=body.name))
    db.commit()
    return {"code": body.code, "name": body.name}


class SectionIn(BaseModel):
    exam_code: str
    key: str
    name: str


@router.post("/sections")
def create_section(body: SectionIn, db: Session = Depends(get_db)) -> dict:
    if db.get(models.Exam, body.exam_code) is None:
        raise HTTPException(404, f"no exam {body.exam_code!r}")
    if db.scalar(select(models.Section).where(
            models.Section.exam_code == body.exam_code, models.Section.key == body.key)):
        raise HTTPException(409, f"section {body.key!r} already exists in {body.exam_code}")
    s = models.Section(exam_code=body.exam_code, key=body.key, name=body.name)
    db.add(s)
    db.commit()
    return {"id": str(s.id), "key": s.key, "name": s.name}


class NodeIn(BaseModel):
    id: str
    exam_code: str
    section_key: str
    name: str
    parent_id: Optional[str] = None


def _section_id(db: Session, exam_code: str, section_key: str):
    s = db.scalar(select(models.Section).where(
        models.Section.exam_code == exam_code, models.Section.key == section_key))
    if s is None:
        raise HTTPException(404, f"no section {section_key!r} in {exam_code}")
    return s.id


@router.post("/topics")
def create_topic(body: NodeIn, db: Session = Depends(get_db)) -> dict:
    if db.get(models.KnowledgeNode, body.id):
        raise HTTPException(409, f"node {body.id!r} already exists")
    sid = _section_id(db, body.exam_code, body.section_key)
    n = models.KnowledgeNode(id=body.id, exam_code=body.exam_code, section_id=sid,
                             kind="topic", name=body.name, parent_id=None)
    db.add(n)
    db.commit()
    return {"id": n.id, "name": n.name, "kind": "topic"}


@router.post("/concepts")
def create_concept(body: NodeIn, db: Session = Depends(get_db)) -> dict:
    if db.get(models.KnowledgeNode, body.id):
        raise HTTPException(409, f"node {body.id!r} already exists")
    sid = _section_id(db, body.exam_code, body.section_key)
    if body.parent_id:
        parent = db.get(models.KnowledgeNode, body.parent_id)
        if parent is None or parent.kind != "topic":
            raise HTTPException(400, "parent_id must be an existing topic node")
    n = models.KnowledgeNode(id=body.id, exam_code=body.exam_code, section_id=sid,
                             kind="concept", name=body.name, parent_id=body.parent_id)
    db.add(n)
    db.commit()
    return {"id": n.id, "name": n.name, "kind": "concept", "parent_id": n.parent_id}


class PrereqIn(BaseModel):
    node_id: str
    prereq_node_id: str


@router.post("/prereqs")
def add_prereq(body: PrereqIn, db: Session = Depends(get_db)) -> dict:
    if db.get(models.KnowledgeNode, body.node_id) is None or \
            db.get(models.KnowledgeNode, body.prereq_node_id) is None:
        raise HTTPException(404, "both node_id and prereq_node_id must exist")
    if body.node_id == body.prereq_node_id:
        raise HTTPException(400, "a node cannot be its own prerequisite")
    if db.scalar(select(models.PrereqEdge).where(
            models.PrereqEdge.node_id == body.node_id,
            models.PrereqEdge.prereq_node_id == body.prereq_node_id)):
        return {"ok": True, "already": True}
    db.add(models.PrereqEdge(node_id=body.node_id, prereq_node_id=body.prereq_node_id))
    db.commit()
    return {"ok": True}


@router.delete("/prereqs")
def del_prereq(body: PrereqIn, db: Session = Depends(get_db)) -> dict:
    row = db.scalar(select(models.PrereqEdge).where(
        models.PrereqEdge.node_id == body.node_id,
        models.PrereqEdge.prereq_node_id == body.prereq_node_id))
    if row:
        db.delete(row)
        db.commit()
    return {"ok": True}


@router.delete("/nodes/{node_id}")
def delete_node(node_id: str, db: Session = Depends(get_db)) -> dict:
    n = db.get(models.KnowledgeNode, node_id)
    if n is None:
        raise HTTPException(404, f"no node {node_id!r}")
    if db.scalar(select(models.KnowledgeNode).where(models.KnowledgeNode.parent_id == node_id)):
        raise HTTPException(409, "node has child concepts — delete or reparent them first")
    if db.scalar(select(models.Item).where(models.Item.concept_node_id == node_id)):
        raise HTTPException(409, "node has items — delete/retire them first")
    db.execute(delete(models.PrereqEdge).where(
        (models.PrereqEdge.node_id == node_id) | (models.PrereqEdge.prereq_node_id == node_id)))
    db.delete(n)
    db.commit()
    return {"ok": True, "deleted": node_id}


# ───────────────────────── items ─────────────────────────
@router.get("/items")
def list_items(exam: Optional[str] = None, section: Optional[str] = None,
               concept: Optional[str] = None, status: Optional[str] = None,
               limit: int = 200, db: Session = Depends(get_db)) -> dict:
    q = select(models.Item)
    if exam:
        q = q.where(models.Item.exam_code == exam)
    if concept:
        q = q.where(models.Item.concept_node_id == concept)
    if status:
        q = q.where(models.Item.status == status)
    items = db.scalars(q.limit(limit)).all()
    sec_key = {s.id: s.key for s in db.scalars(select(models.Section)).all()}
    out = []
    for it in items:
        if section and sec_key.get(it.section_id) != section:
            continue
        out.append({
            "item_id": it.item_id, "version": it.version, "exam_code": it.exam_code,
            "section": sec_key.get(it.section_id), "concept_node_id": it.concept_node_id,
            "difficulty_d": it.difficulty_d, "format": it.format, "num_options": it.num_options,
            "stem": it.stem, "options": it.options, "correct_answer": it.correct_answer,
            "solution": it.solution, "status": it.status,
        })
    return {"count": len(out), "items": out}


@router.post("/items", response_model=IngestReport)
def create_item(body: ItemIn, db: Session = Depends(get_db)) -> IngestReport:
    """Create (or upsert) a single item. Goes through the same validated ingest path as bulk upload,
    so the authored-vs-derived boundary (no IRT a/b/c authoring) is enforced identically."""
    return ingestion.ingest_items(db, [body])


class ItemPatch(BaseModel):
    model_config = ConfigDict(extra="forbid")
    stem: Optional[str] = None
    options: Optional[list[str]] = None
    correct_answer: Optional[str] = None
    solution: Optional[str] = None
    difficulty_d: Optional[int] = Field(default=None, ge=-2, le=2)
    status: Optional[str] = None
    negative_marking: Optional[bool] = None


@router.patch("/items/{item_id}")
def edit_item(item_id: str, body: ItemPatch, db: Session = Depends(get_db)) -> dict:
    it = db.get(models.Item, item_id)
    if it is None:
        raise HTTPException(404, f"no item {item_id!r}")
    data = body.model_dump(exclude_unset=True)
    for k, v in data.items():
        setattr(it, k, v)
    if data:
        it.version = (it.version or 1) + 1
    db.commit()
    return {"ok": True, "item_id": item_id, "version": it.version}


@router.post("/items/{item_id}/approve")
def approve_item(item_id: str, db: Session = Depends(get_db)) -> dict:
    it = db.get(models.Item, item_id)
    if it is None:
        raise HTTPException(404, f"no item {item_id!r}")
    it.status = "approved"
    db.commit()
    return {"ok": True, "item_id": item_id, "status": "approved"}


@router.post("/items/{item_id}/retire")
def retire_item(item_id: str, db: Session = Depends(get_db)) -> dict:
    it = db.get(models.Item, item_id)
    if it is None:
        raise HTTPException(404, f"no item {item_id!r}")
    it.status = "retired"
    db.commit()
    return {"ok": True, "item_id": item_id, "status": "retired"}


@router.delete("/items/{item_id}")
def delete_item(item_id: str, db: Session = Depends(get_db)) -> dict:
    it = db.get(models.Item, item_id)
    if it is None:
        raise HTTPException(404, f"no item {item_id!r}")
    db.delete(it)
    db.commit()
    return {"ok": True, "deleted": item_id}


@router.post("/items/upload-xlsx")
async def upload_xlsx(file: UploadFile = File(...), db: Session = Depends(get_db)) -> dict:
    """Admin-gated bulk upload — same question-bank workbook format as the authors' path: one sheet
    per exam (tab name = exam) OR a single sheet with an `Exam` column. Builds the graph from
    Topic/Subtopic/Prerequisites and ingests every question atomically; returns the import report."""
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover
        raise HTTPException(500, "openpyxl not installed")
    data = await file.read()
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"could not read workbook: {e}")
    rows: list[dict] = []
    for ws in wb.worksheets:
        it = ws.iter_rows(values_only=True)
        try:
            header = [str(h).strip() if h is not None else "" for h in next(it)]
        except StopIteration:
            continue
        for r in it:
            if r is None or all(c is None for c in r):
                continue
            row: dict[str, Any] = dict(zip(header, r))
            if not row.get("Exam"):
                row["Exam"] = ws.title
            rows.append(row)
    if not rows:
        raise HTTPException(400, "no data rows found in any sheet")
    return question_bank.import_question_bank(db, rows)


# ───────────────────────── admin management ─────────────────────────
@router.get("/admins")
def list_admins(db: Session = Depends(get_db)) -> list[dict]:
    rows = db.scalars(select(models.AdminUser)).all()
    out = []
    for r in rows:
        acc = db.get(models.Account, r.account_id)
        out.append({"account_id": str(r.account_id), "role": r.role,
                    "email": acc.email if acc else None,
                    "display_name": acc.display_name if acc else None})
    return out


class GrantIn(BaseModel):
    email: str


@router.post("/admins")
def grant(body: GrantIn, db: Session = Depends(get_db)) -> dict:
    acc = grant_admin(db, body.email)
    return {"ok": True, "account_id": str(acc.id), "email": acc.email}


@router.delete("/admins/{account_id}")
def revoke(account_id: str, admin: models.Account = Depends(require_admin),
           db: Session = Depends(get_db)) -> dict:
    if str(admin.id) == account_id:
        raise HTTPException(400, "you cannot revoke your own admin access")
    ok = revoke_admin(db, account_id)
    if not ok:
        raise HTTPException(404, "that account is not an admin")
    return {"ok": True, "revoked": account_id}
