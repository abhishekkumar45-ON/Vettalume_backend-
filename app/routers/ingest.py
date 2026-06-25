from __future__ import annotations

import io

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy.orm import Session

from ..deps import get_current_learner, get_db  # noqa: F401 (learner dep reserved for authz later)
from ..schemas import IngestReport, ItemIn
from ..services.admin_auth import require_admin
from ..services.ingestion import ingest_items
from ..services.question_bank import import_question_bank

router = APIRouter(prefix="/ingest", tags=["ingest"], dependencies=[Depends(require_admin)])  # content upload is admin-only

# Excel template columns (order-independent; header row required):
XLSX_COLUMNS = [
    "item_id", "exam_code", "section_key", "concept_node_id", "archetype_id", "grid_cell",
    "difficulty_d", "format", "num_options", "negative_marking", "stem", "options",
    "correct_answer", "solution", "time_benchmark_s", "usage_scope",
]


@router.post("/items", response_model=IngestReport)
def ingest_json(rows: list[ItemIn], db: Session = Depends(get_db)) -> IngestReport:
    """Bulk-ingest items as JSON. Validate-all-then-commit; rejects the whole batch on any error."""
    return ingest_items(db, rows)


@router.post("/items/xlsx", response_model=IngestReport)
async def ingest_xlsx(file: UploadFile = File(...), db: Session = Depends(get_db)) -> IngestReport:
    """Bulk-ingest items from an .xlsx authored in your generator pipeline.
    `options` cell is pipe-delimited, e.g.  6|7|8|9 ."""
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover
        raise HTTPException(500, "openpyxl not installed")

    data = await file.read()
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        raise HTTPException(400, "empty workbook")

    parsed: list[ItemIn] = []
    for r in rows_iter:
        if r is None or all(c is None for c in r):
            continue
        rowdict = dict(zip(header, r))
        try:
            parsed.append(_row_to_item(rowdict))
        except Exception as e:
            raise HTTPException(400, f"row parse error for item_id={rowdict.get('item_id')}: {e}")

    return ingest_items(db, parsed)


@router.post("/question-bank/xlsx")
async def ingest_question_bank_xlsx(file: UploadFile = File(...), db: Session = Depends(get_db)) -> dict:
    """Import a Vettalume question-bank workbook (.xlsx) — the bulk-upload path for authors.

    Handles both layouts:
      * one sheet with an `Exam` column, and
      * one sheet PER exam (the tab name is the exam, e.g. CAT / GMAT / GRE).

    Builds the knowledge graph from Topic / Subtopic / Prerequisites, resolves A–E / multi / sequence
    answers, and ingests the questions across all sheets in one atomic transaction. `IRT b` is ignored
    (it is derived by calibration). Dangling prerequisites are reported as warnings, not failures.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover
        raise HTTPException(500, "openpyxl not installed")

    data = await file.read()
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)

    rows: list[dict] = []
    for ws in wb.worksheets:
        it = ws.iter_rows(values_only=True)
        try:
            header = [str(h).strip() if h is not None else "" for h in next(it)]
        except StopIteration:
            continue  # empty sheet
        for r in it:
            if r is None or all(c is None for c in r):
                continue
            row = dict(zip(header, r))
            if not row.get("Exam"):           # no Exam column -> the tab name IS the exam
                row["Exam"] = ws.title
            rows.append(row)

    if not rows:
        raise HTTPException(400, "no data rows found in any sheet")

    return import_question_bank(db, rows)
    def s(key):
        v = d.get(key)
        return None if v is None or str(v).strip() == "" else str(v).strip()

    options = s("options")
    payload = {
        "item_id": s("item_id"),
        "exam_code": s("exam_code"),
        "section_key": s("section_key"),
        "concept_node_id": s("concept_node_id"),
        "archetype_id": s("archetype_id"),
        "grid_cell": s("grid_cell"),
        "difficulty_d": int(d["difficulty_d"]),
        "format": (s("format") or "mcq").lower(),
        "num_options": int(d.get("num_options") or 4),
        "negative_marking": str(d.get("negative_marking") or "").strip().lower() in ("1", "true", "yes"),
        "stem": s("stem"),
        "options": [o.strip() for o in options.split("|")] if options else None,
        "correct_answer": s("correct_answer"),
        "solution": s("solution"),
        "time_benchmark_s": int(d["time_benchmark_s"]) if s("time_benchmark_s") else None,
        "usage_scope": (s("usage_scope") or "both").lower(),
    }
    return ItemIn(**{k: v for k, v in payload.items() if v is not None})
