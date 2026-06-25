#!/usr/bin/env python3
"""Interactive IRT mock — an adaptive mock that shows your ability (theta) after every answer.

It serves the highest-information question at your current ability (a real CAT move), records each
answer in the COLD `full_mock` context (the only context IRT scores), then re-estimates theta by EAP
on the -3..+3 scale and prints it with its SE and 95% interval. Each run is a fresh session from
theta = 0. Items are uncalibrated, so b comes from the authored difficulty, c from the option count,
a = 1 — exactly how a cold-start diagnostic scores before calibration.

Usage:
    python scripts/irt_mock.py [PATH_TO_BANK.xlsx] [--exam GMAT] [--section Quant] [--max 25]

If no path is given it looks for the file you uploaded in ~/Downloads.
"""
import argparse
import os
import random
import sys

os.environ.setdefault("SERVE_ONLY_APPROVED", "false")        # the GMAT bank is mostly draft
os.environ.setdefault("DATABASE_URL", "sqlite+pysqlite:///:memory:")
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl  # noqa: E402
from sqlalchemy import create_engine, select  # noqa: E402
from sqlalchemy.orm import Session  # noqa: E402
from sqlalchemy.pool import StaticPool  # noqa: E402

from app import models  # noqa: E402
from app.services import calibration, irt  # noqa: E402
from app.services.question_bank import import_question_bank  # noqa: E402
from app.services.state import record_response  # noqa: E402

LETTERS = "ABCDEF"


def load_rows(path: str, exam: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    if exam not in wb.sheetnames:
        sys.exit(f"sheet '{exam}' not found. Sheets in file: {wb.sheetnames}")
    ws = wb[exam]
    hdr = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        d = dict(zip(hdr, r))
        d["Exam"] = exam
        rows.append(d)
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("bank", nargs="?",
                    default=os.path.expanduser("~/Downloads/Vettalume_Question_Bank_by_CAT_GMAT.xlsx"))
    ap.add_argument("--exam", default="GMAT")
    ap.add_argument("--section", default=None, help="e.g. Quant, 'Data Insights', Verbal (default: all)")
    ap.add_argument("--max", type=int, default=25)
    args = ap.parse_args()

    if not os.path.exists(args.bank):
        sys.exit(f"bank file not found:\n  {args.bank}\n"
                 f"pass the path explicitly:  python scripts/irt_mock.py /path/to/bank.xlsx")

    eng = create_engine(os.environ["DATABASE_URL"],
                        connect_args={"check_same_thread": False}, poolclass=StaticPool)
    models.Base.metadata.create_all(eng)
    db = Session(eng, autoflush=False)

    rep = import_question_bank(db, load_rows(args.bank, args.exam))
    learner = models.Account(email="mock@local", display_name="mock")
    db.add(learner); db.flush()

    items = list(db.scalars(select(models.Item).where(models.Item.exam_code == args.exam)).all())
    if args.section:
        sec_ids = [s.id for s in db.scalars(
            select(models.Section).where(models.Section.exam_code == args.exam)).all()
            if args.section.lower() in (s.key.lower(), s.name.lower())]
        items = [it for it in items if it.section_id in sec_ids]
    if not items:
        sys.exit(f"no items found for {args.exam}{' / ' + args.section if args.section else ''}")

    sid = f"mock-{random.randint(1000, 9999)}"
    served: set[str] = set()
    answered: list[tuple] = []           # (a, b, c, u)
    theta = 0.0

    ins = rep.get("questions", {}).get("inserted", len(items))
    print(f"\n=== IRT mock — {args.exam}{' / ' + args.section if args.section else ''} "
          f"({len(items)} items in pool, {ins} ingested) ===")
    print("Answer with the option letter. 's' = skip, 'q' = quit.\n")

    n = 0
    while n < args.max:
        pool = [it for it in items if it.item_id not in served]
        if not pool:
            print("Pool exhausted — no more unseen questions.")
            break

        # max-information selection at the current ability
        def info(it):
            a, b, c = calibration.active_params(db, it)
            return irt.information_3pl(theta, a, b, c)
        item = max(pool, key=info)
        a, b, c = calibration.active_params(db, item)
        served.add(item.item_id)

        opts = item.options or []
        print(f"--- Q{n+1}  [{item.item_id}]   difficulty b = {b:+.0f}   (c={c:.2f})")
        print(f"  {item.stem}")
        for i, opt in enumerate(opts):
            print(f"    {LETTERS[i]}. {opt}")
        try:
            ans = input("  Your answer: ").strip().upper()
        except EOFError:
            break
        if ans == "Q":
            break
        if ans == "S" or ans == "":
            print("  (skipped)\n")
            continue

        chosen = opts[LETTERS.index(ans)] if (len(ans) == 1 and ans in LETTERS[:len(opts)]) else ans
        _resp, correct, _state = record_response(
            db, learner, item, context="full_mock", answer_given=chosen, correct=None,
            response_time_ms=None, attempt_number=1, hints_used=0, session_id=sid)
        answered.append((a, b, c, 1 if correct else 0))
        theta, se = irt.eap_ability(answered)
        n += 1

        mark = "\u2713 correct" if correct else "\u2717 wrong  "
        if se == float("inf"):
            print(f"  {mark}   theta = {theta:+.3f}   SE —        (answered {n})\n")
        else:
            print(f"  {mark}   theta = {theta:+.3f}   SE {se:4.2f}   "
                  f"95% CI [{theta - 2 * se:+.2f}, {theta + 2 * se:+.2f}]   (answered {n})\n")

    if answered:
        theta, se = irt.eap_ability(answered)
        ncorrect = sum(u for *_, u in answered)
        tail = f"   SE {se:.2f}" if se != float("inf") else ""
        print(f"=== Final ability: theta = {theta:+.3f}{tail}   |   {ncorrect}/{len(answered)} correct ===")
        print("    (theta is on the -3..+3 IRT scale; 0 = average. Run /irt/calibrate later to replace "
              "the authored-difficulty priors with data-driven a, b, c.)")
    else:
        print("No questions answered.")


if __name__ == "__main__":
    main()
